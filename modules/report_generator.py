"""
Report Generator Module
Excel report generation for DQ results
"""

import pandas as pd
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List
from .config import AppConfig
from .utils import clean_value


class ExcelReportGenerator:
    """Generate comprehensive Excel reports for DQ results"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_report(
        self,
        results_df: pd.DataFrame,
        overall_score: float,
        column_scores: Dict[str, float],
        dimension_scores: Dict[str, float],
        rulebook: Dict,
        master_data_columns: List[str],
        output_filename: str = None
    ) -> Path:
        """
        Generate comprehensive Excel report
        
        Returns:
            Path to generated Excel file
        """
        if output_filename is None:
            from .utils import get_timestamp
            output_filename = f"DQ_Report_{get_timestamp()}.xlsx"
        
        output_path = AppConfig.OUTPUT_DIR / "DQ_Reports" / output_filename
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # 1. DQ Score Summary
        self._add_summary_sheet(wb, overall_score, column_scores, dimension_scores)
        
        # 2. Full Results
        self._add_results_sheet(wb, results_df)
        
        # 3. Column-wise Annexures
        self._add_column_annexures(wb, results_df, master_data_columns)
        
        # 4. Dimension Analysis
        self._add_dimension_analysis(wb, dimension_scores)
        
        # 5. Rulebook Reference
        self._add_rulebook_sheet(wb, rulebook)
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    def _add_summary_sheet(self, wb: Workbook, overall: float, col_scores: Dict, dim_scores: Dict):
        """Add DQ Score Summary sheet"""
        ws = wb.create_sheet("DQ Score Summary", 0)
        
        # Overall score
        ws.append(["Data Quality Score Summary"])
        ws.append([])
        ws.append(["Overall DQ Score", f"{overall:.2f}%"])
        ws.append([])
        
        # Column scores
        ws.append(["Column-wise DQ Scores"])
        ws.append(["Column", "DQ Score (%)", "Status"])
        
        for col, score in sorted(col_scores.items(), key=lambda x: x[1]):
            status = "PASSED" if score == 100 else "FAILED"
            ws.append([col, f"{score:.2f}", status])
        
        ws.append([])
        
        # Dimension scores
        if dim_scores:
            ws.append(["Dimension-wise DQ Scores"])
            ws.append(["Dimension", "DQ Score (%)", "Status"])
            
            for dim, score in dim_scores.items():
                status = "PASSED" if score == 100 else "FAILED"
                ws.append([dim, f"{score:.2f}", status])
        
        # Format
        self._format_sheet(ws)
    
    def _add_results_sheet(self, wb: Workbook, results_df: pd.DataFrame):
        """Add full results sheet"""
        ws = wb.create_sheet("DQ Results", 1)
        
        # Get display columns
        display_cols = [c for c in results_df.columns if not c.startswith("_")]
        results_display = results_df[display_cols].copy()
        
        # Clean values
        for col in results_display.columns:
            results_display[col] = results_display[col].apply(clean_value)
        
        # Write to sheet
        for r_idx, row in enumerate(dataframe_to_rows(results_display, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Header formatting
                if r_idx == 1:
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                
                cell.border = self.border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_column_annexures(self, wb: Workbook, results_df: pd.DataFrame, columns: List[str]):
        """Add column-wise annexure sheets"""
        for col in columns:
            # Skip metadata columns
            if col.startswith("_") or col in [
                "Issues", "Count of issues", "Failed_Rules", 
                "Failed_Columns", "Issue categories"
            ]:
                continue
            
            # Get records with issues in this column
            issues_in_col = results_df[
                results_df["_failed_columns_list"].apply(
                    lambda x: col in x if isinstance(x, list) else False
                )
            ]
            
            if len(issues_in_col) == 0:
                continue
            
            # Create sheet (max 31 chars for Excel)
            sheet_name = f"Ann_{col}"[:31]
            ws = wb.create_sheet(sheet_name)
            
            # Select relevant columns
            display_cols = [c for c in results_df.columns if not c.startswith("_")]
            annexure_data = issues_in_col[display_cols].copy()
            
            # Clean values
            for c in annexure_data.columns:
                annexure_data[c] = annexure_data[c].apply(clean_value)
            
            # Write to sheet
            for r_idx, row in enumerate(dataframe_to_rows(annexure_data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    if r_idx == 1:
                        cell.fill = self.header_fill
                        cell.font = self.header_font
                    
                    cell.border = self.border
            
            # Format
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_dimension_analysis(self, wb: Workbook, dim_scores: Dict[str, float]):
        """Add dimension analysis sheet"""
        if not dim_scores:
            return
        
        ws = wb.create_sheet("Dimension Analysis")
        
        ws.append(["Dimension", "DQ Score (%)", "Status"])
        
        for dim, score in dim_scores.items():
            status = "PASSED" if score == 100 else "FAILED"
            ws.append([dim, f"{score:.2f}", status])
        
        self._format_sheet(ws)
    
    def _add_rulebook_sheet(self, wb: Workbook, rulebook: Dict):
        """Add rulebook reference sheet"""
        ws = wb.create_sheet("Rulebook")
        
        ws.append(["Rule Configuration"])
        ws.append([])
        
        # Metadata
        if "metadata" in rulebook:
            ws.append(["Metadata"])
            for key, value in rulebook["metadata"].items():
                ws.append([key, str(value)])
            ws.append([])
        
        # Rules
        ws.append(["Rules"])
        ws.append(["Column", "Rule Type", "Dimension", "Message", "Expression", "Severity"])
        
        for rule in rulebook.get("rules", []):
            ws.append([
                rule.get("column", ""),
                rule.get("rule_type", ""),
                rule.get("dimension", ""),
                rule.get("message", ""),
                str(rule.get("expression", "")),
                rule.get("severity", "")
            ])
        
        self._format_sheet(ws)
    
    def _format_sheet(self, ws):
        """Apply basic formatting to sheet"""
        for row in ws.iter_rows():
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Format first row as header
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.fill = self.header_fill
                cell.font = self.header_font
