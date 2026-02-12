# ── stdlib ─────────────────────────────────────────────────────────────────
import traceback
import datetime
from io import BytesIO

# ── third-party ────────────────────────────────────────────────────────────
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Wedge

# ── DQ Engine modules ──────────────────────────────────────────────────────
from modules.config           import AppConfig
from modules.file_loader      import FileLoaderService
from modules.rulebook_builder import RulebookBuilderService
from modules.rule_executor    import RuleExecutorEngine
from modules.scoring_engine   import ScoringService
from modules.report_generator import ExcelReportGenerator
from modules.ui_components    import UIComponents
from modules.utils            import (
    setup_directories, save_uploaded_file, clean_temp_directory,
)

# ── DataMaturity modules ───────────────────────────────────────────────────
from DataMaturity.config import (
    UNIQU_PURPLE, UNIQU_MAGENTA, UNIQU_LAVENDER,
    UNIQU_LIGHT_BG, UNIQU_TEXT, UNIQU_GREY,
    RATING_LABELS, RATING_TO_SCORE,
    DEFAULT_MASTER_OBJECTS, MATURITY_DIMS, QUESTION_BANK,
)

from DataMaturity.helpers import (
    dq_score_to_maturity_level,
    init_maturity_state,
    build_question_df,
    sync_response_tables,
    autofill_dq_dimension,
    compute_all_scores,
    validate_responses,
    to_excel_bytes,
)

from DataMaturity.visualizations   import render_slide_png
from DataMaturity.report_generator import build_pdf_bytes


# ══════════════════════════════════════════════════════════════════════════
#  APP CONFIG & EXTERNAL CSS
# ══════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Enterprise DQ & Maturity Platform",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

setup_directories()


def load_css():
    """Load external stylesheet from assets folder."""
    try:
        with open("assets/styles.css", encoding="utf-8") as f:
            st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
    except FileNotFoundError:
        st.warning("⚠️ styles.css not found in assets/ folder")


# ══════════════════════════════════════════════════════════════════════════
#  SESSION STATE INITIALIZATION
# ══════════════════════════════════════════════════════════════════════════
def _init_state() -> None:
    # Navigation
    if "page" not in st.session_state:
        st.session_state["page"] = "home"

    # DQ results (populated after DQ run; consumed by Maturity)
    for key, default in {
        "dq_score":       None,
        "dq_dim_scores":  None,
        "dq_results_df":  None,
        "dq_object_name": "Customer",
        "dq_excel_path":  None,
    }.items():
        if key not in st.session_state:
            st.session_state[key] = default

    # Maturity state (uses keys defined in DataMaturity/helpers.py)
    init_maturity_state()
    
    # Policy Hub state
    if "policies" not in st.session_state:
        st.session_state["policies"] = []
    
    # Case Management state
    if "cases" not in st.session_state:
        st.session_state["cases"] = []


# ══════════════════════════════════════════════════════════════════════════
#  UTILITY FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════
def get_timestamp_filename(prefix: str, extension: str) -> str:
    """Generate filename with current date and time."""
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.{extension}"


# ══════════════════════════════════════════════════════════════════════════
#  VISUALIZATION FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════
def _gauge_png(score: float) -> bytes:
    """Semicircle gauge for overall DQ score with enhanced design."""
    fig, ax = plt.subplots(figsize=(5, 3.2), dpi=150)
    fig.patch.set_facecolor('#fafafa')
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 0.65)
    ax.axis("off")

    # Background arc
    ax.add_patch(Wedge((0.5, 0.05), 0.40, 0, 180, width=0.12,
                       facecolor="#e5e7eb", edgecolor="white", lw=3))
    
    # Score arc with gradient color
    ang = score / 100 * 180
    if score >= 80:
        col = "#10b981"
    elif score >= 60:
        col = "#f59e0b"
    else:
        col = "#ef4444"
    
    ax.add_patch(Wedge((0.5, 0.05), 0.40, 0, ang, width=0.12,
                       facecolor=col, edgecolor="white", lw=3))

    # Score text
    ax.text(0.5, 0.32, f"{score:.1f}%", ha="center", va="center",
            fontsize=28, fontweight="bold", color="#6d28d9", family="sans-serif")
    ax.text(0.5, 0.18, "Overall DQ Score", ha="center", va="center",
            fontsize=11, color="#57534e", family="sans-serif", weight=600)

    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", pad_inches=0.15, facecolor='#fafafa')
    plt.close(fig)
    return buf.getvalue()


def _dim_bar_png(dim_scores: dict) -> bytes | None:
    """Horizontal bar chart – DQ dimension scores with enhanced styling."""
    if not dim_scores:
        return None

    dims   = list(dim_scores.keys())
    scores = [dim_scores[d] for d in dims]
    
    # Enhanced color scheme
    cols = []
    for s in scores:
        if s >= 80:
            cols.append("#10b981")
        elif s >= 60:
            cols.append("#f59e0b")
        else:
            cols.append("#ef4444")

    fig, ax = plt.subplots(figsize=(8, max(3, len(dims) * 0.8)), dpi=140)
    fig.patch.set_facecolor('#fafafa')
    ax.set_facecolor('#ffffff')

    bars = ax.barh(dims, scores, color=cols, height=0.6, edgecolor="white", linewidth=2)
    ax.set_xlim(0, 112)
    ax.set_xlabel("DQ Score (%)", color="#1c1917", fontsize=11, weight=600, family="sans-serif")
    ax.tick_params(colors="#44403c", labelsize=10)
    ax.spines[["top", "right", "bottom"]].set_visible(False)
    ax.spines["left"].set_color("#d6d3d1")
    ax.spines["left"].set_linewidth(1.5)

    # Reference lines
    ax.axvline(80, color="#6d28d9",  lw=1.5, ls="--", alpha=0.6, label="Excellent (80%)")
    ax.axvline(60, color="#7c3aed", lw=1.5, ls=":",  alpha=0.6, label="Good (60%)")
    ax.legend(fontsize=9, loc="lower right", frameon=True, fancybox=True, shadow=True)

    # Value labels
    for bar, sc in zip(bars, scores):
        ax.text(bar.get_width() + 2, bar.get_y() + bar.get_height() / 2,
                f"{sc:.1f}%", va="center", fontsize=10,
                fontweight="bold", color="#1c1917", family="sans-serif")

    plt.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor='#fafafa')
    plt.close(fig)
    return buf.getvalue()


def _mat_bar_png(dim_vals: dict) -> bytes | None:
    """Horizontal bar chart – maturity dimension scores with light blue theme."""
    if not dim_vals:
        return None

    dims   = list(dim_vals.keys())
    scores = [dim_vals[d] for d in dims]
    
    # Light blue color scheme
    cols = []
    for s in scores:
        if s >= 4:
            cols.append("#0369a1")  # Dark blue
        elif s >= 3:
            cols.append("#0ea5e9")  # Sky blue
        else:
            cols.append("#7dd3fc")  # Light blue

    fig, ax = plt.subplots(figsize=(10, max(3, len(dims) * 0.9)), dpi=140)
    fig.patch.set_facecolor('#f0f9ff')
    ax.set_facecolor('#ffffff')

    bars = ax.barh(dims, scores, color=cols, height=0.6, edgecolor="white", linewidth=2)
    ax.set_xlim(0, 6.0)
    ax.set_xlabel("Maturity Score (1 = Adhoc  →  5 = Optimised)",
                  color="#0c4a6e", fontsize=11, weight=600, family="sans-serif")

    # Reference lines
    ax.axvline(3.0, color="#38bdf8", lw=1.5, ls="--", alpha=0.7, label="Defined (3)")
    ax.axvline(4.0, color="#0284c7", lw=1.5, ls="--", alpha=0.7, label="Managed (4)")
    ax.legend(fontsize=9, loc="lower right", frameon=True, fancybox=True, shadow=True)

    ax.tick_params(colors="#0c4a6e", labelsize=10)
    ax.spines[["top", "right", "bottom"]].set_visible(False)
    ax.spines["left"].set_color("#bae6fd")
    ax.spines["left"].set_linewidth(2)

    # Value labels
    for bar, sc in zip(bars, scores):
        ax.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height() / 2,
                f"{sc:.2f}", va="center", fontsize=11,
                fontweight="bold", color="#0c4a6e", family="sans-serif")

    plt.tight_layout()
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor='#f0f9ff')
    plt.close(fig)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════
#  COMBINED EXCEL (DQ + Maturity)
# ══════════════════════════════════════════════════════════════════════════
def _combined_excel(dq_score: float, dq_dim_scores: dict | None, mat_excel: bytes) -> bytes:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = load_workbook(BytesIO(mat_excel))

    # Sheet 1: DQ Score Summary with styling
    ws_dq = wb.create_sheet("DQ Score Summary", 0)
    
    # Header styling
    header_fill = PatternFill(start_color="6d28d9", end_color="7c3aed", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    ws_dq.append(["Metric", "Value"])
    for cell in ws_dq[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws_dq.append(["Overall DQ Score (%)",  f"{dq_score:.1f}%"])
    ws_dq.append(["Mapped Maturity Level", dq_score_to_maturity_level(dq_score)])

    if dq_dim_scores:
        for dim, sc in dq_dim_scores.items():
            ws_dq.append([f"DQ – {dim}", f"{sc:.1f}%"])

    # Column widths
    ws_dq.column_dimensions['A'].width = 30
    ws_dq.column_dimensions['B'].width = 20

    # Sheet 2: DQ Results (first 1000 rows)
    dq_df = st.session_state.get("dq_results_df")
    if dq_df is not None:
        display_cols = [c for c in dq_df.columns if not c.startswith("_")]
        ws_res = wb.create_sheet("DQ Results", 1)
        ws_res.append(display_cols)
        
        # Style header
        for cell in ws_res[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for _, row in dq_df[display_cols].head(1000).iterrows():
            ws_res.append([str(v) if v is not None else "" for v in row.tolist()])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ══════════════════════════════════════════════════════════════════════════
#  PAGE: HOME - Premium Home Page Design
# ══════════════════════════════════════════════════════════════════════════
def page_home():
    # Animated background effect
    st.markdown('<div class="animated-bg"></div>', unsafe_allow_html=True)
    
    # Hero Section
    st.markdown('<div class="home-hero">', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("# 📊 Enterprise DQ & Maturity Platform")
        st.markdown(
            "**One integrated platform** for comprehensive Data Quality validation, "
            "DAMA Maturity assessment, Policy Management, and Case Tracking"
        )
    st.markdown('</div>', unsafe_allow_html=True)

    st.divider()

    # DQ Completion Banner
    if st.session_state.dq_score is not None:
        sc  = st.session_state.dq_score
        lvl = dq_score_to_maturity_level(sc)
        st.markdown('<div class="banner success">', unsafe_allow_html=True)
        col1, col2 = st.columns([3, 1])
        with col1:
            st.markdown(
                f"✅ **DQ Assessment Completed**  \n"
                f"**Score:** {sc:.1f}% | **Level:** {lvl} | **Object:** {st.session_state.dq_object_name}"
            )
        with col2:
            if st.button("View Results →", use_container_width=True, key="home_to_dq"):
                st.session_state["page"] = "dq"
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
        st.divider()

    # Feature Cards Grid
    st.markdown('<h2 style="text-align: center; margin-bottom: 2.5rem; font-size: 2.2rem;">Integrated Solutions</h2>', unsafe_allow_html=True)
    
    # Row 1: DQ and Maturity
    col1, col_space1, col2 = st.columns([1, 0.05, 1])
    
    with col1:
        st.markdown(
            '''
            <div class="feature-card">
                <div class="feature-card-icon">🔍</div>
                <h3>Data Quality Assessment</h3>
                <p>Upload your dataset and rules configuration to get comprehensive DQ scores, dimension-wise breakdowns, column-level analyses, and automated maturity mapping.</p>
                <div style="margin-top: 1.5rem;">
            </div>
            </div>
            ''',
            unsafe_allow_html=True,
        )
        if st.button("Start DQ Assessment →", use_container_width=True, key="home_dq_btn"):
            st.session_state["page"] = "dq"
            st.rerun()

    with col2:
        st.markdown(
            '''
            <div class="feature-card">
                <div class="feature-card-icon">📈</div>
                <h3>Data Maturity Assessment</h3>
                <p>Answer DAMA questionnaire to assess organizational maturity. Get slide-style visuals, professional PDF reports, Excel exports, and optional auto-population from DQ scores.</p>
                <div style="margin-top: 1.5rem;">
            </div>
            </div>
            ''',
            unsafe_allow_html=True,
        )
        if st.button("Start Maturity Assessment →", use_container_width=True, key="home_mat_btn"):
            st.session_state["page"] = "maturity"
            st.rerun()

    st.markdown('<div style="height: 1.5rem;"></div>', unsafe_allow_html=True)
    
    # Row 2: Policy Hub and Case Management
    col3, col_space2, col4 = st.columns([1, 0.05, 1])
    
    with col3:
        st.markdown(
            '''
            <div class="feature-card">
                <div class="feature-card-icon">📋</div>
                <h3>Policy Hub</h3>
                <p>Centralized repository for data governance policies. Create, manage, version, and track compliance with organizational data policies and standards.</p>
                <div style="margin-top: 1.5rem;">
            </div>
            </div>
            ''',
            unsafe_allow_html=True,
        )
        if st.button("Open Policy Hub →", use_container_width=True, key="home_policy_btn"):
            st.session_state["page"] = "policy"
            st.rerun()

    with col4:
        st.markdown(
            '''
            <div class="feature-card">
                <div class="feature-card-icon">🎯</div>
                <h3>Case Management System</h3>
                <p>Track and resolve data quality issues, assign ownership, monitor progress, and maintain complete audit trails for all data-related cases and incidents.</p>
                <div style="margin-top: 1.5rem;">
            </div>
            </div>
            ''',
            unsafe_allow_html=True,
        )
        if st.button("Open Case Management →", use_container_width=True, key="home_case_btn"):
            st.session_state["page"] = "cases"
            st.rerun()

    st.divider()

    # Mapping Table
    st.markdown("#### DQ Score to Maturity Level Mapping")
    
    mapping_data = [
        ("95% – 100%", "⭐ Optimised", "Continuous improvement; proactive governance"),
        ("80% – 94%", "✅ Managed", "Monitored, measured, formalised roles"),
        ("60% – 79%", "🔵 Defined", "Standardised and consistently followed"),
        ("40% – 59%", "🟡 Repeatable", "Some processes defined but inconsistent"),
        ("0% – 39%", "🔴 Adhoc", "Unstructured, reactive, varies widely"),
    ]
    
    for score, level, desc in mapping_data:
        st.markdown(
            f'<div class="mapping-row"><div class="mapping-score">{score}</div><div class="mapping-emoji">{level}</div><div>{desc}</div></div>',
            unsafe_allow_html=True
        )


# ══════════════════════════════════════════════════════════════════════════
#  PAGE: DQ ASSESSMENT - Enhanced DQ Page
# ══════════════════════════════════════════════════════════════════════════
def page_dq():
    # Enhanced Sidebar Navigation
    with st.sidebar:
        st.markdown("### 🧭 Navigation")
        nav_col1, nav_col2 = st.columns(2)
        with nav_col1:
            if st.button("🏠 Home", use_container_width=True, key="dq_home"):
                st.session_state["page"] = "home"
                st.rerun()
        with nav_col2:
            if st.button("📈 Maturity", use_container_width=True, key="dq_maturity"):
                st.session_state["page"] = "maturity"
                st.rerun()
        
        nav_col3, nav_col4 = st.columns(2)
        with nav_col3:
            if st.button("📋 Policies", use_container_width=True, key="dq_policy"):
                st.session_state["page"] = "policy"
                st.rerun()
        with nav_col4:
            if st.button("🎯 Cases", use_container_width=True, key="dq_cases"):
                st.session_state["page"] = "cases"
                st.rerun()
        
        st.divider()
        UIComponents.render_sidebar()

    # Header
    st.markdown("# 🔍 Data Quality Assessment")
    st.markdown(
        "Upload your master dataset and rules configuration to generate "
        "comprehensive DQ reports with detailed scoring and analysis."
    )
    st.divider()

    # File Upload Section
    st.markdown("### 📁 Input Files")
    col1, col2 = st.columns(2)
    
    with col1:
        data_file = st.file_uploader(
            "Master Dataset",
            type=AppConfig.SUPPORTED_DATA_FORMATS,
            help="CSV, Excel, or JSON format",
            key="dq_data_uploader"
        )
    
    with col2:
        rules_file = st.file_uploader(
            "Rules Configuration",
            type=AppConfig.SUPPORTED_RULES_FORMATS + ["json"],
            help="CSV, Excel rules sheet, or JSON rulebook",
            key="dq_rules_uploader"
        )

    if not data_file or not rules_file:
        st.markdown('<div class="info-box">', unsafe_allow_html=True)
        st.markdown("👉 Please upload both a dataset and rules configuration to begin.")
        st.markdown('</div>', unsafe_allow_html=True)
        UIComponents.render_welcome_screen()
        return

    # Configuration Section
    st.markdown("### ⚙️ Configuration")
    col1, col2 = st.columns([2, 1])
    
    with col1:
        obj_name = st.text_input(
            "Master Data Object Name",
            value=st.session_state.get("dq_object_name", "Customer"),
            placeholder="e.g., Customer, Vendor, Product",
        )
    
    with col2:
        sheet_name = None
        if data_file.name.lower().endswith((".xlsx", ".xls", ".xlsm")):
            loader = FileLoaderService()
            tmp    = AppConfig.TEMP_DIR / data_file.name
            tmp.write_bytes(data_file.getbuffer())
            sheets = loader.get_sheet_names(tmp)
            if len(sheets) > 1:
                sheet_name = st.selectbox("Select Sheet", sheets, key="dq_sheet")

    UIComponents.render_file_format_help()
    st.divider()

    # Run Button
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        run_button = st.button(
            "🚀 Run DQ Assessment",
            type="primary",
            use_container_width=True,
            key="dq_run"
        )

    if not run_button:
        return

    try:
        clean_temp_directory()
        pb   = st.progress(0, text="📂 Saving files...")
        stat = st.empty()

        stat.text("📂 Saving files…");       pb.progress(5, text="📂 Saving files...")
        data_path  = save_uploaded_file(data_file,  AppConfig.TEMP_DIR)
        rules_path = save_uploaded_file(rules_file, AppConfig.TEMP_DIR)

        stat.text("📊 Loading dataset…");    pb.progress(15, text="📊 Loading dataset...")
        loader = FileLoaderService()
        df     = loader.load_dataframe(data_path, sheet_name=sheet_name)
        cols   = list(df.columns)
        st.info(f"✅ Loaded **{len(df):,}** records · **{len(cols)}** columns")

        stat.text("🔧 Building rulebook…");  pb.progress(30, text="🔧 Building rulebook...")
        rb_svc = RulebookBuilderService()
        if rules_file.name.lower().endswith(".json"):
            rulebook = rb_svc.load_json_rulebook(rules_path)
        else:
            rulebook = rb_svc.build_from_rules_dataset(
                loader.load_dataframe(rules_path), cols)

        stat.text("✅ Executing rules…");    pb.progress(50, text="✅ Executing rules...")
        executor = RuleExecutorEngine(df, rulebook)
        results  = executor.execute_all_rules()
        combos   = executor.get_combination_duplicates()

        stat.text("📊 Scoring…");            pb.progress(70, text="📊 Calculating scores...")
        scorer     = ScoringService()
        overall    = scorer.calculate_overall_score(results)
        col_scores = scorer.calculate_column_scores(results, cols)
        dim_scores = scorer.calculate_dimension_scores(results)

        stat.text("💾 Generating Excel report…");      pb.progress(85, text="💾 Generating Excel report...")
        rgen = ExcelReportGenerator(
            results_df=results,
            rulebook=rulebook,
            all_columns=cols,
            column_scores=col_scores,
            overall_score=overall,
            dimension_scores=dim_scores,
            duplicate_combinations=combos,
        )
        
        # Generate filename with timestamp
        excel_filename = get_timestamp_filename(f"DQ_Report_{obj_name or 'Dataset'}", "xlsx")
        xl_path = AppConfig.OUTPUT_DIR / excel_filename
        rgen.generate_report(AppConfig.OUTPUT_DIR, filename=excel_filename)
        rb_path = rgen.save_rulebook_json(AppConfig.OUTPUT_DIR, rulebook)
        
        pb.progress(100, text="✅ Complete!")
        stat.success("✅ Assessment completed successfully!")

        # Save to session
        st.session_state["dq_score"]       = overall
        st.session_state["dq_dim_scores"]  = dim_scores
        st.session_state["dq_results_df"]  = results
        st.session_state["dq_object_name"] = obj_name or "Customer"
        st.session_state["dq_excel_path"]  = xl_path

        # Sync to maturity
        st.session_state["mat_objects"] = [obj_name] if obj_name else DEFAULT_MASTER_OBJECTS[:]
        autofill_dq_dimension(overall)

        st.markdown('<div class="banner success">', unsafe_allow_html=True)
        st.markdown("✅ **DQ Assessment Completed Successfully!**")
        st.markdown('</div>', unsafe_allow_html=True)
        st.divider()

        # ── Results Dashboard ──────────────────────────────────────────
        st.markdown("## 📊 Results Dashboard")
        g1, g2 = st.columns([1, 2])
        with g1:
            st.image(_gauge_png(overall), use_container_width=True)
        with g2:
            bar = _dim_bar_png(dim_scores)
            if bar:
                st.image(bar, use_container_width=True)

        st.divider()
        UIComponents.render_results_dashboard(overall, results, col_scores, dim_scores)
        st.divider()
        
        # Enhanced Download Section with timestamped filenames
        st.markdown("### 📥 Download Reports")
        d1, d2, d3 = st.columns(3)
        
        with d1:
            with open(xl_path, "rb") as f:
                st.download_button(
                    "📊 DQ Excel Report",
                    data=f.read(),
                    file_name=excel_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
        
        with d2:
            if rb_path and rb_path.exists():
                with open(rb_path, "rb") as f:
                    rb_filename = get_timestamp_filename("Rulebook", "json")
                    st.download_button(
                        "📋 Rulebook JSON",
                        data=f.read(),
                        file_name=rb_filename,
                        mime="application/json",
                        use_container_width=True,
                    )
        
        with d3:
            st.info(f"✅ {len(cols)} columns analyzed")
        
        st.divider()
        UIComponents.render_detailed_views(rulebook, results, col_scores, dim_scores)

        st.divider()
        lvl = dq_score_to_maturity_level(overall)
        st.markdown(
            f'<div class="banner">'
            f'💡 DQ Score <span class="purple-text">{overall:.1f}%</span> maps to maturity level '
            f'<span class="purple-text">{lvl}</span>. '
            f'This has been auto-filled in the Data Quality assessment dimension.'
            f'</div>',
            unsafe_allow_html=True
        )

        st.divider()
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            if st.button("📈 Continue to Maturity Assessment →", type="primary", use_container_width=True, key="dq_to_mat"):
                st.session_state["page"] = "maturity"
                st.rerun()

    except Exception as e:
        st.markdown('<div class="banner danger">', unsafe_allow_html=True)
        st.markdown(f"❌ **Error:** {e}")
        st.markdown('</div>', unsafe_allow_html=True)
        with st.expander("🔍 Technical Details"):
            st.code(traceback.format_exc())


def _apply_editor_edits(dim: str, editor_key: str) -> None:
    """Apply editor changes to maturity responses."""
    widget_state = st.session_state.get(editor_key)
    if not widget_state:
        return

    edited_rows = widget_state.get("edited_rows", {})
    if not edited_rows:
        return

    df = st.session_state.mat_responses[dim].copy()
    for row_idx, changes in edited_rows.items():
        for col, val in changes.items():
            df.at[int(row_idx), col] = val

    st.session_state.mat_responses[dim] = df


# ══════════════════════════════════════════════════════════════════════════
#  PAGE: MATURITY ASSESSMENT - Enhanced with Light Blue Theme
# ══════════════════════════════════════════════════════════════════════════
def _do_submit() -> None:
    objects   = st.session_state.mat_objects
    dims      = st.session_state.mat_dims
    responses = st.session_state.mat_responses
    cn        = st.session_state.mat_client_name or "Client"
    bm        = float(st.session_state.mat_benchmark)
    tg        = float(st.session_state.mat_target)
    lt        = float(st.session_state.mat_low_thr)
    dq_score  = st.session_state.get("dq_score")

    ok, msg = validate_responses(responses, dims, objects)
    if not ok:
        st.error(f"⚠️ Validation failed: {msg}")
        return

    with st.spinner("⚙️ Computing scores and building reports…"):
        dim_table, overall = compute_all_scores(objects, dims, responses)

        domain_display = {
            dim: float(np.nanmean(dim_table.loc[dim].values))
            for dim in dims
        }
        exec_score = float(np.nanmean(overall.values)) if len(overall) else 0.0

        slide_png = render_slide_png(
            client_name=cn,
            domain_scores=domain_display,
            exec_score=exec_score if np.isfinite(exec_score) else 0.0,
            benchmark=bm,
            target=tg,
        )

        pdf_bytes = build_pdf_bytes(
            client_name=cn,
            slide_png=slide_png,
            dim_table=dim_table,
            overall=overall,
            detail_tables=responses,
            dq_score=dq_score,
        )

        mat_excel = to_excel_bytes(
            dim_table=dim_table,
            overall=overall,
            detail_tables=responses,
            low_thr=lt,
            objects=objects,
        )

        if dq_score is not None:
            combined_excel = _combined_excel(
                dq_score,
                st.session_state.get("dq_dim_scores"),
                mat_excel,
            )
        else:
            combined_excel = mat_excel

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    st.session_state["mat_submitted"] = True
    st.session_state["mat_payload"]   = {
        "dim_table":      dim_table,
        "overall":        overall,
        "slide_png":      slide_png,
        "mat_excel":      mat_excel,
        "combined_excel": combined_excel,
        "pdf_bytes":      pdf_bytes,
        "client_name":    cn,
        "ts":             ts,
    }
    st.rerun()


def page_maturity():
    dq_score  = st.session_state.get("dq_score")
    submitted = st.session_state.get("mat_submitted", False)

    # Enhanced Sidebar
    with st.sidebar:
        st.markdown("### 🧭 Navigation")
        nav_col1, nav_col2 = st.columns(2)
        with nav_col1:
            if st.button("🏠 Home", use_container_width=True, key="mat_home"):
                st.session_state["page"] = "home"
                st.rerun()
        with nav_col2:
            if st.button("🔍 DQ", use_container_width=True, key="mat_dq"):
                st.session_state["page"] = "dq"
                st.rerun()
        
        nav_col3, nav_col4 = st.columns(2)
        with nav_col3:
            if st.button("📋 Policies", use_container_width=True, key="mat_policy"):
                st.session_state["page"] = "policy"
                st.rerun()
        with nav_col4:
            if st.button("🎯 Cases", use_container_width=True, key="mat_cases"):
                st.session_state["page"] = "cases"
                st.rerun()
        
        st.divider()

        st.markdown("### ⚙️ Configuration")
        st.session_state["mat_client_name"] = st.text_input(
            "Client Name",
            value=st.session_state.get("mat_client_name", ""),
            placeholder="Organisation name",
            disabled=submitted,
        )

        all_obj_opts = list(dict.fromkeys(
            DEFAULT_MASTER_OBJECTS + st.session_state.mat_objects
        ))
        st.session_state["mat_objects"] = st.multiselect(
            "Master Data Objects",
            options=all_obj_opts,
            default=st.session_state.mat_objects,
            disabled=submitted,
        )

        st.session_state["mat_dims"] = st.multiselect(
            "Maturity Dimensions",
            options=MATURITY_DIMS,
            default=st.session_state.mat_dims,
            disabled=submitted,
        )

        st.divider()
        st.markdown("### 📊 Thresholds")
        st.session_state["mat_low_thr"] = st.slider(
            "Exception threshold (≤)", 1.0, 5.0,
            float(st.session_state.get("mat_low_thr", 2.0)), 0.5,
            disabled=submitted,
        )

        st.divider()
        st.markdown("### 🎯 Benchmark / Target")
        st.session_state["mat_benchmark"] = st.number_input(
            "Industry Benchmark", 1.0, 5.0,
            float(st.session_state.get("mat_benchmark", 3.0)), 0.1,
            disabled=submitted,
        )
        st.session_state["mat_target"] = st.number_input(
            "Target Score", 1.0, 5.0,
            float(st.session_state.get("mat_target", 3.0)), 0.1,
            disabled=submitted,
        )

    if not st.session_state.mat_objects or not st.session_state.mat_dims:
        st.markdown('<div class="info-box">', unsafe_allow_html=True)
        st.markdown("👉 Please select at least one **Object** and one **Dimension** in the sidebar.")
        st.markdown('</div>', unsafe_allow_html=True)
        st.stop()

    prev_objs = st.session_state.get("_last_sync_objects")
    prev_dims = st.session_state.get("_last_sync_dims")
    curr_objs = st.session_state.mat_objects
    curr_dims = st.session_state.mat_dims

    if prev_objs != curr_objs or prev_dims != curr_dims:
        sync_response_tables()
        for d in curr_dims:
            st.session_state.pop(f"mat_snap_{d}", None)
        st.session_state["_last_sync_objects"] = list(curr_objs)
        st.session_state["_last_sync_dims"]    = list(curr_dims)

    if dq_score is not None and not st.session_state.get("dq_autofilled"):
        autofill_dq_dimension(dq_score)
        st.session_state.pop("mat_snap_Data Quality", None)
        st.session_state["dq_autofilled"] = True

    # ── REPORT VIEW (after submit) ─────────────────────────────────────
    if submitted and st.session_state.get("mat_payload"):
        p  = st.session_state["mat_payload"]
        cn = p["client_name"]
        ts = p["ts"]

        st.markdown("# ✅ Data Maturity Assessment Report")

        if dq_score is not None:
            lvl = dq_score_to_maturity_level(dq_score)
            st.markdown(
                f'<div class="banner success">'
                f'**DQ Engine Score:** {dq_score:.1f}% → **Level:** {lvl} (applied to Data Quality dimension)'
                f'</div>',
                unsafe_allow_html=True
            )

        st.markdown("### 📊 Summary Slide")
        st.image(p["slide_png"], use_container_width=True)
        st.divider()

        # Light blue themed tables
        st.markdown("""
            <style>
            .dataframe tbody tr:hover {
                background-color: rgba(224, 242, 254, 0.5) !important;
            }
            .dataframe thead th {
                background: linear-gradient(135deg, #e0f2fe 0%, #bae6fd 100%) !important;
                color: #0c4a6e !important;
            }
            </style>
        """, unsafe_allow_html=True)

        t1, t2 = st.columns(2)
        with t1:
            st.markdown("#### Dimension-wise Maturity")
            
            # Apply light blue gradient styling
            styled_dim = p["dim_table"].style\
                .format("{:.2f}")\
                .background_gradient(cmap="Blues", axis=None, vmin=1, vmax=5)
            
            st.dataframe(styled_dim, use_container_width=True)
            
        with t2:
            st.markdown("#### Overall Maturity Score")
            
            styled_overall = pd.DataFrame(p["overall"]).T.style\
                .format("{:.2f}")\
                .background_gradient(cmap="Blues", axis=None, vmin=1, vmax=5)
            
            st.dataframe(styled_overall, use_container_width=True)

        st.divider()
        st.markdown("#### Scores by Dimension")
        dim_vals = {
            dim: float(np.nanmean(p["dim_table"].loc[dim].values))
            for dim in p["dim_table"].index
        }
        bar_img = _mat_bar_png(dim_vals)
        if bar_img:
            st.image(bar_img, use_container_width=True)

        st.divider()
        st.markdown("### 📥 Download Reports")
        safe_cn = cn.replace(" ", "_")
        d1, d2, d3 = st.columns(3)

        with d1:
            pdf_filename = get_timestamp_filename(f"Maturity_Report_{safe_cn}", "pdf")
            st.download_button(
                "📄 PDF Report",
                data=p["pdf_bytes"],
                file_name=pdf_filename,
                mime="application/pdf",
                use_container_width=True,
            )

        with d2:
            mat_excel_filename = get_timestamp_filename(f"Maturity_Assessment_{safe_cn}", "xlsx")
            st.download_button(
                "📊 Maturity Excel",
                data=p["mat_excel"],
                file_name=mat_excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with d3:
            combined_filename = get_timestamp_filename(f"DQ_Maturity_Combined_{safe_cn}", "xlsx")
            st.download_button(
                "🔗 Combined Excel",
                data=p["combined_excel"],
                file_name=combined_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        st.divider()
        if st.button("✏️ Edit Responses", use_container_width=True, key="mat_edit"):
            st.session_state["mat_submitted"] = False
            st.session_state["mat_payload"]   = {}
            st.rerun()

        st.stop()

    # ── QUESTIONNAIRE VIEW ─────────────────────────────────────────────
    st.markdown("# 📈 Data Maturity Assessment")

    if dq_score is not None:
        lvl = dq_score_to_maturity_level(dq_score)
        st.markdown(
            f'<div class="banner success">'
            f'✅ DQ Score **{dq_score:.1f}%** → level **{lvl}** auto-filled in *Data Quality* dimension.'
            f'</div>',
            unsafe_allow_html=True
        )
    else:
        st.markdown(
            '<div class="banner">'
            '💡 Run the **DQ Assessment** to automatically populate the *Data Quality* dimension.'
            '</div>',
            unsafe_allow_html=True
        )

    st.markdown("Rate each question per master data object, then submit for analysis.")
    st.divider()

    dims = st.session_state.mat_dims
    tabs = st.tabs(dims)

    for i, dim in enumerate(dims):
        with tabs[i]:
            st.markdown(f"### {dim}")

            if dim == "Data Quality" and dq_score is not None:
                lvl = dq_score_to_maturity_level(dq_score)
                st.markdown(
                    f'<div class="banner">'
                    f'Auto-populated from DQ Score **{dq_score:.1f}%** → **{lvl}**. '
                    f'You can adjust individual ratings as needed.'
                    f'</div>',
                    unsafe_allow_html=True
                )

            cfg = {
                "Weight": st.column_config.NumberColumn(
                    "Weight", min_value=0.0, step=0.5),
            }
            for obj in st.session_state.mat_objects:
                cfg[obj] = st.column_config.SelectboxColumn(
                    obj, options=RATING_LABELS, required=True)

            editor_key = f"mat_editor_{dim}"

            st.data_editor(
                st.session_state.mat_responses[dim],
                use_container_width=True,
                hide_index=True,
                column_config=cfg,
                disabled=["Question ID", "Section", "Question"],
                key=editor_key,
                on_change=_apply_editor_edits,
                args=(dim, editor_key),
            )

    st.divider()
    col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
    with col2:
        if st.button("🚀 Submit & Generate Report", type="primary", use_container_width=True, key="mat_submit"):
            _do_submit()
    with col3:
        st.info("**Submit** to generate visuals and downloadable reports.")


# ══════════════════════════════════════════════════════════════════════════
#  PAGE: POLICY HUB
# ══════════════════════════════════════════════════════════════════════════
def page_policy_hub():
    with st.sidebar:
        st.markdown("### 🧭 Navigation")
        nav_col1, nav_col2 = st.columns(2)
        with nav_col1:
            if st.button("🏠 Home", use_container_width=True, key="policy_home"):
                st.session_state["page"] = "home"
                st.rerun()
        with nav_col2:
            if st.button("🔍 DQ", use_container_width=True, key="policy_dq"):
                st.session_state["page"] = "dq"
                st.rerun()
        
        nav_col3, nav_col4 = st.columns(2)
        with nav_col3:
            if st.button("📈 Maturity", use_container_width=True, key="policy_maturity"):
                st.session_state["page"] = "maturity"
                st.rerun()
        with nav_col4:
            if st.button("🎯 Cases", use_container_width=True, key="policy_cases"):
                st.session_state["page"] = "cases"
                st.rerun()
        
        st.divider()
        st.markdown("### 📋 Policy Statistics")
        st.metric("Total Policies", len(st.session_state.policies))
        if st.session_state.policies:
            active = sum(1 for p in st.session_state.policies if p.get("status") == "Active")
            st.metric("Active Policies", active)

    st.markdown("# 📋 Policy Hub")
    st.markdown("Centralized repository for data governance policies and standards.")
    st.divider()

    # Add New Policy
    with st.expander("➕ Create New Policy", expanded=False):
        with st.form("new_policy_form"):
            col1, col2 = st.columns(2)
            with col1:
                policy_name = st.text_input("Policy Name", placeholder="e.g., Data Retention Policy")
                policy_category = st.selectbox(
                    "Category",
                    ["Data Quality", "Data Privacy", "Data Security", "Data Governance", "Compliance", "Other"]
                )
            with col2:
                policy_owner = st.text_input("Policy Owner", placeholder="e.g., Data Governance Team")
                policy_status = st.selectbox("Status", ["Draft", "Active", "Under Review", "Archived"])
            
            policy_description = st.text_area("Policy Description", height=100)
            
            col3, col4 = st.columns(2)
            with col3:
                effective_date = st.date_input("Effective Date")
            with col4:
                review_date = st.date_input("Next Review Date")
            
            submitted = st.form_submit_button("Create Policy", type="primary", use_container_width=True)
            
            if submitted and policy_name:
                new_policy = {
                    "id": len(st.session_state.policies) + 1,
                    "name": policy_name,
                    "category": policy_category,
                    "owner": policy_owner,
                    "status": policy_status,
                    "description": policy_description,
                    "effective_date": effective_date.strftime("%Y-%m-%d"),
                    "review_date": review_date.strftime("%Y-%m-%d"),
                    "created_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "version": "1.0"
                }
                st.session_state.policies.append(new_policy)
                st.success(f"✅ Policy '{policy_name}' created successfully!")
                st.rerun()

    st.divider()

    # Display Policies
    if st.session_state.policies:
        st.markdown("### 📚 Policy Repository")
        
        # Filter controls
        col1, col2, col3 = st.columns(3)
        with col1:
            filter_category = st.selectbox(
                "Filter by Category",
                ["All"] + ["Data Quality", "Data Privacy", "Data Security", "Data Governance", "Compliance", "Other"]
            )
        with col2:
            filter_status = st.selectbox("Filter by Status", ["All", "Draft", "Active", "Under Review", "Archived"])
        with col3:
            search_term = st.text_input("Search Policies", placeholder="Search by name...")
        
        # Filter policies
        filtered_policies = st.session_state.policies
        if filter_category != "All":
            filtered_policies = [p for p in filtered_policies if p.get("category") == filter_category]
        if filter_status != "All":
            filtered_policies = [p for p in filtered_policies if p.get("status") == filter_status]
        if search_term:
            filtered_policies = [p for p in filtered_policies if search_term.lower() in p.get("name", "").lower()]
        
        # Display policies as cards
        for policy in filtered_policies:
            with st.container():
                st.markdown(f"""
                    <div style="background: white; padding: 1.5rem; border-radius: 0.75rem; 
                                margin-bottom: 1rem; border-left: 4px solid #6d28d9; 
                                box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                        <h3 style="color: #6d28d9; margin-bottom: 0.5rem;">{policy['name']}</h3>
                        <p style="color: #57534e; margin-bottom: 1rem;">{policy.get('description', 'No description')}</p>
                        <div style="display: flex; gap: 1rem; flex-wrap: wrap;">
                            <span style="background: #f0f9ff; color: #0369a1; padding: 0.25rem 0.75rem; 
                                        border-radius: 0.375rem; font-size: 0.875rem;">
                                📁 {policy.get('category', 'N/A')}
                            </span>
                            <span style="background: {'#d1fae5' if policy.get('status') == 'Active' else '#fef3c7'}; 
                                        color: {'#065f46' if policy.get('status') == 'Active' else '#92400e'}; 
                                        padding: 0.25rem 0.75rem; border-radius: 0.375rem; font-size: 0.875rem;">
                                {policy.get('status', 'N/A')}
                            </span>
                            <span style="color: #78716c; font-size: 0.875rem;">
                                👤 {policy.get('owner', 'N/A')}
                            </span>
                            <span style="color: #78716c; font-size: 0.875rem;">
                                📅 Effective: {policy.get('effective_date', 'N/A')}
                            </span>
                            <span style="color: #78716c; font-size: 0.875rem;">
                                🔄 Version: {policy.get('version', '1.0')}
                            </span>
                        </div>
                    </div>
                """, unsafe_allow_html=True)
        
        # Export policies
        st.divider()
        if st.button("📥 Export All Policies to Excel", use_container_width=True):
            df_policies = pd.DataFrame(st.session_state.policies)
            excel_data = BytesIO()
            with pd.ExcelWriter(excel_data, engine='openpyxl') as writer:
                df_policies.to_excel(writer, sheet_name='Policies', index=False)
            excel_data.seek(0)
            
            filename = get_timestamp_filename("Policy_Repository", "xlsx")
            st.download_button(
                "Download Excel",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    else:
        st.markdown('<div class="info-box">', unsafe_allow_html=True)
        st.markdown("📋 No policies created yet. Use the form above to create your first policy.")
        st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
#  PAGE: CASE MANAGEMENT
# ══════════════════════════════════════════════════════════════════════════
def page_case_management():
    with st.sidebar:
        st.markdown("### 🧭 Navigation")
        nav_col1, nav_col2 = st.columns(2)
        with nav_col1:
            if st.button("🏠 Home", use_container_width=True, key="case_home"):
                st.session_state["page"] = "home"
                st.rerun()
        with nav_col2:
            if st.button("🔍 DQ", use_container_width=True, key="case_dq"):
                st.session_state["page"] = "dq"
                st.rerun()
        
        nav_col3, nav_col4 = st.columns(2)
        with nav_col3:
            if st.button("📈 Maturity", use_container_width=True, key="case_maturity"):
                st.session_state["page"] = "maturity"
                st.rerun()
        with nav_col4:
            if st.button("📋 Policies", use_container_width=True, key="case_policy"):
                st.session_state["page"] = "policy"
                st.rerun()
        
        st.divider()
        st.markdown("### 🎯 Case Statistics")
        st.metric("Total Cases", len(st.session_state.cases))
        if st.session_state.cases:
            open_cases = sum(1 for c in st.session_state.cases if c.get("status") in ["Open", "In Progress"])
            st.metric("Open Cases", open_cases)

    st.markdown("# 🎯 Case Management System")
    st.markdown("Track and resolve data quality issues with complete audit trails.")
    st.divider()

    # Create New Case
    with st.expander("➕ Create New Case", expanded=False):
        with st.form("new_case_form"):
            col1, col2 = st.columns(2)
            with col1:
                case_title = st.text_input("Case Title", placeholder="e.g., Invalid Customer Email Addresses")
                case_priority = st.selectbox("Priority", ["Low", "Medium", "High", "Critical"])
            with col2:
                case_category = st.selectbox(
                    "Category",
                    ["Data Quality Issue", "Data Breach", "Compliance Violation", "System Error", "Other"]
                )
                case_assigned_to = st.text_input("Assigned To", placeholder="e.g., John Doe")
            
            case_description = st.text_area("Case Description", height=100)
            
            col3, col4 = st.columns(2)
            with col3:
                case_source = st.selectbox("Source", ["DQ Assessment", "Manual Report", "Automated Alert", "Customer Complaint"])
            with col4:
                case_due_date = st.date_input("Due Date")
            
            submitted = st.form_submit_button("Create Case", type="primary", use_container_width=True)
            
            if submitted and case_title:
                new_case = {
                    "id": f"CASE-{len(st.session_state.cases) + 1:04d}",
                    "title": case_title,
                    "description": case_description,
                    "priority": case_priority,
                    "category": case_category,
                    "status": "Open",
                    "assigned_to": case_assigned_to,
                    "source": case_source,
                    "created_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "due_date": case_due_date.strftime("%Y-%m-%d"),
                    "updated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "comments": []
                }
                st.session_state.cases.append(new_case)
                st.success(f"✅ Case '{new_case['id']}' created successfully!")
                st.rerun()

    st.divider()

    # Display Cases
    if st.session_state.cases:
        st.markdown("### 📋 Case List")
        
        # Filter controls
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            filter_status = st.selectbox(
                "Filter by Status",
                ["All", "Open", "In Progress", "Resolved", "Closed"]
            )
        with col2:
            filter_priority = st.selectbox(
                "Filter by Priority",
                ["All", "Low", "Medium", "High", "Critical"]
            )
        with col3:
            filter_category = st.selectbox(
                "Filter by Category",
                ["All", "Data Quality Issue", "Data Breach", "Compliance Violation", "System Error", "Other"]
            )
        with col4:
            search_term = st.text_input("Search Cases", placeholder="Search by title...")
        
        # Filter cases
        filtered_cases = st.session_state.cases
        if filter_status != "All":
            filtered_cases = [c for c in filtered_cases if c.get("status") == filter_status]
        if filter_priority != "All":
            filtered_cases = [c for c in filtered_cases if c.get("priority") == filter_priority]
        if filter_category != "All":
            filtered_cases = [c for c in filtered_cases if c.get("category") == filter_category]
        if search_term:
            filtered_cases = [c for c in filtered_cases if search_term.lower() in c.get("title", "").lower()]
        
        # Display cases as cards
        for case in filtered_cases:
            priority_colors = {
                "Low": "#d1fae5",
                "Medium": "#fef3c7",
                "High": "#fed7aa",
                "Critical": "#fecaca"
            }
            priority_text_colors = {
                "Low": "#065f46",
                "Medium": "#92400e",
                "High": "#9a3412",
                "Critical": "#991b1b"
            }
            
            with st.container():
                st.markdown(f"""
                    <div style="background: white; padding: 1.5rem; border-radius: 0.75rem; 
                                margin-bottom: 1rem; border-left: 4px solid #6d28d9; 
                                box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                        <div style="display: flex; justify-content: space-between; align-items: start; margin-bottom: 0.5rem;">
                            <h3 style="color: #6d28d9; margin: 0;">{case['id']}: {case['title']}</h3>
                            <span style="background: {priority_colors.get(case.get('priority'), '#f3f4f6')}; 
                                        color: {priority_text_colors.get(case.get('priority'), '#1f2937')}; 
                                        padding: 0.25rem 0.75rem; border-radius: 0.375rem; font-size: 0.875rem; 
                                        font-weight: 600;">
                                {case.get('priority', 'N/A')} Priority
                            </span>
                        </div>
                        <p style="color: #57534e; margin-bottom: 1rem;">{case.get('description', 'No description')}</p>
                        <div style="display: flex; gap: 1rem; flex-wrap: wrap;">
                            <span style="background: #f0f9ff; color: #0369a1; padding: 0.25rem 0.75rem; 
                                        border-radius: 0.375rem; font-size: 0.875rem;">
                                📁 {case.get('category', 'N/A')}
                            </span>
                            <span style="background: #fef3c7; color: #92400e; padding: 0.25rem 0.75rem; 
                                        border-radius: 0.375rem; font-size: 0.875rem;">
                                {case.get('status', 'N/A')}
                            </span>
                            <span style="color: #78716c; font-size: 0.875rem;">
                                👤 {case.get('assigned_to', 'Unassigned')}
                            </span>
                            <span style="color: #78716c; font-size: 0.875rem;">
                                📅 Due: {case.get('due_date', 'N/A')}
                            </span>
                            <span style="color: #78716c; font-size: 0.875rem;">
                                🕒 Created: {case.get('created_at', 'N/A')}
                            </span>
                        </div>
                    </div>
                """, unsafe_allow_html=True)
        
        # Export cases
        st.divider()
        if st.button("📥 Export All Cases to Excel", use_container_width=True):
            df_cases = pd.DataFrame(st.session_state.cases)
            # Remove comments column for cleaner export
            if 'comments' in df_cases.columns:
                df_cases = df_cases.drop('comments', axis=1)
            
            excel_data = BytesIO()
            with pd.ExcelWriter(excel_data, engine='openpyxl') as writer:
                df_cases.to_excel(writer, sheet_name='Cases', index=False)
            excel_data.seek(0)
            
            filename = get_timestamp_filename("Case_Management", "xlsx")
            st.download_button(
                "Download Excel",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    else:
        st.markdown('<div class="info-box">', unsafe_allow_html=True)
        st.markdown("🎯 No cases created yet. Use the form above to create your first case.")
        st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
#  START APPLICATION
# ══════════════════════════════════════════════════════════════════════════
load_css()
_init_state()

{
    "home":     page_home,
    "dq":       page_dq,
    "maturity": page_maturity,
    "policy":   page_policy_hub,
    "cases":    page_case_management,
}[st.session_state.page]()
